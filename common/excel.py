import json
import os

from openpyxl import load_workbook


def execl_to_json(execl):
    try:
        wb = load_workbook(execl, data_only=True)
        json_sheet = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            json_row = {}
            for r in range(5, 48):
                json_col = []
                for c in range(1, 21):
                    json_col.append(ws.cell(r, c).value)
                json_row[ws.cell(r, 1).value] = json_col
            json_sheet[sheet] = json_row
        return json.dumps(json_sheet)
    except BaseException as msg:
        print(msg)
        return None


def excel_save(excel, order_id):
    try:
        load_workbook(excel)
        with open(f'excel/{excel}', 'wb+') as f:
            for chunk in excel.chunks():
                f.write(chunk)
            f.close()
        return f'{order_id}.xlsx'
    except BaseException as msg:
        print(msg)
        return None


def excel_remove(order_id):
    try:
        excel_path = os.getcwd() + f'excel\\{order_id}.xlsx'
        os.remove(excel_path)
    except BaseException as msg:
        print(msg)
        return None
